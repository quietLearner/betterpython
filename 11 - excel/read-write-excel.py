import openpyxl
import os
import pathlib


def read_excel_file(file_name):
    # read excel file
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active

    # read data from excel file
    for row in range(0, sheet.max_row):
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[row].value is not None:
                print(f"{col}", col[row].value)


def main() -> None:
    print(pathlib.Path(__file__).parent.resolve())
    print(os.getcwd())
    read_excel_file(f"{pathlib.Path(__file__).parent.resolve()}\\a.xlsx")


if __name__ == "__main__":
    main()
